from openpyxl import load_workbook

fn = '/home/sibears/Desktop/Work_With_Excel/test.xlsx'
wb = load_workbook(fn)

ws= wb['data'] # лист  с которым работаем

print(f'Employees[A1]={ws["A5"].value}') #чтение по А*
print(f'data = {ws.cell(row=6, column=1).value}')# чтение по строке и столбцу
ws.append(["daniil sib_coder", "anna"])#добавление по порядку

ws['A5'] ="Hello World" #точечьное добавление
ws['A6'] ="Hello Rus"
ws['B2'] ="Hello World"

wb.save(fn)
wb.close
#самый удобный гайд https://pythobyte.com/openpyxl-python-read-write-excel-files-daa58214/
